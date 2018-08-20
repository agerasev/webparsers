#!/usr/bin/python3

import sys
import json
from urllib.request import Request, urlopen
import openpyxl
import xlsxwriter


headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0"}


if len(sys.argv) < 3:
	print("usage: python.exe ./fetch.py ((<infile.xlsx>) | (<word0> [word1] [word2] ... [wordN])) <outfile.xlsx>")

words = []
if sys.argv[1].endswith(".xlsx"):
	ifn = sys.argv[1]
	iwb = openpyxl.load_workbook(ifn)
	iws = iwb[iwb.get_sheet_names()[0]]

	k = 0
	while True:
		word = iws.cell(row=(1 + k), column=1).value
		if word is None:
			break;
		word = str(word).strip()
		words.append(word)
		k += 1
else:
	for w in sys.argv[1:-1]:
		words.append(w)

ofn = sys.argv[-1]
owb = xlsxwriter.Workbook(ofn)
ows = owb.add_worksheet()

encoding = "utf-8"
for i, word in enumerate(words):
	try:
		print("get '%s'" % word)
		req = Request(
			"http://synonymonline.ru/assets/json/synonyms.json",
			data=("word=%s" % word).encode(encoding),
			headers=headers
		)
		res = urlopen(req).read().decode(encoding)
		res = json.loads(res)

		ows.write(i, 0, res["word"])
		ows.write(i, 1, res["word_baseform"])
		ows.write(i, 2, ", ".join(res["synonyms"]))
	except:
		print("error")

owb.close()
