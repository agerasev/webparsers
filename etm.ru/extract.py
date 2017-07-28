#!/usr/bin/python3

import sys
import re
import json
from urllib.request import Request, urlopen
from urllib.parse   import quote
import openpyxl
import xlsxwriter


def rmtag(s):
	return re.sub("<[^>]*>", "", s).strip()

def cookie_store(d):
	s = ""
	for k, v in d.items():
		s += k + "=" + v + ";"
	return s

def cookie_load(s):
	s = re.sub("expires=[^;]*;", "", s)
	d = {}
	for kv in s.strip().split(","):
		(k, v) = tuple(kv.strip().split(";")[0].strip().split("="))
		d[k] = v
	return d


headers = { "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0" }
cookies = {}

if len(sys.argv) < 2 or len(sys.argv[1]) == 0:
	print("usage: python.exe extract.py <C:/output/directory> <C:/source/file.xlsx>")
	exit()

outdir = sys.argv[1]
srcfn = "requests.xlsm"
if len(sys.argv) >= 3 and len(sys.argv[2]) > 0:
	srcfn = sys.argv[2]

print("opening '%s' file..." % srcfn)
wb = openpyxl.load_workbook(srcfn)
ws = wb[wb.get_sheet_names()[0]]

login = ws.cell(row=2, column=2).value
password = ""
anon = False
if login is not None:
	login = str(login)
	print("login: " + login)
	password = ws.cell(row=2, column=3).value
	if password is None:
		password = ""
	else:
		password = str(password)
else:
	print("login field is empty - using anonymous mode")
	login = ""
	anon = True


if anon:
	print("entering site...")
else:
	print("logging into site...")

requests = [
	"http://nsk.etm.ru/cat/catalog.html",
	"https://nsk.etm.ru/ns2000/auth_block.php?retPath=/cat/catalog.html",
	"https://nsk.etm.ru/ns2000/data-login.php?t=login&log=" + login + "&pwd=" + password + "&org-id=78096&smsCode=&session=&new_pwd=&new_pwd2=&_=1499165067733"
]

def make_request(req):
	req = Request(
		req,
		headers=headers
	)
	sc = urlopen(req).getheader("Set-Cookie")
	#print(sc)
	cookies.update(cookie_load(sc))
	headers["Cookie"] = cookie_store(cookies)

if anon:
	make_request(requests[0])
else:
	for req in requests:
		make_request(req)
		#print(cookies)


print("extracting searches...")

fields = [
	("Код товара", 0),
	("Наименование", 2),
	("№ Партии / Артикул", 4),
	("Производитель", 5),
	("Упак.", 6),
	("Ед.изм.", 13),
	("Цена", 10),
	("Цена розничная", 11),
]

headers["Referer"] = "http://nsk.etm.ru/cat/catalog.html";
headers["Cookie"] = cookie_store(cookies)

def extract(res, page):
	data = json.loads(res)
	if int(data["page"]) < page:
		return []
	print("page: " + data["page"] + ", rows: " + str(len(data["rows"])))

	ret = []
	for r in data["rows"]:
		c = r["cell"]
		cell = []
		for f in fields:
			 cell.append(rmtag(c[f[1]]))
		#cell.append(kw)
		ret.append(cell)
	return ret

def fetch_logged(kw, page, rows=40):
	req = Request(
		"http://nsk.etm.ru/cat/data-catalog.html?" +
		"st=16&dst=70100&rg=" + quote("УРНВС") + "&login_type=iPRO&" +
		"org-id=690032659&prc348=false&city=&cena=361&cls=&val=" + quote(kw) +
		"&type=text&my_cat=0&_search=false&nd=1499165073462&" +
		"rows=" + str(rows) + "&page=" + str(page) +
		"&sidx=cls81+asc,+store&sord=desc",
		headers=headers
	)
	res = urlopen(req).read().decode("windows-1251")
	return extract(res, page)
	

def fetch_anon(kw, page, rows=40):
	req = Request(
		"http://nsk.etm.ru/cat/data-catalog.html?" + 
		"st=0&dst=0&rg=&login_type=CoolHacker&org-id=78096&prc348=false&city=" + quote("ЭТ") + "&" +
		"cena=98&val=" + quote(kw) + "&" +
		"cls=&type=text&_search=false&nd=1499004246844&rows=" + str(rows) + "&page=" + str(page) + "&" +
		"sidx=info&sord=desc",
		headers=headers
	)
	res = urlopen(req).read().decode("windows-1251")
	return extract(res, page)

k = 0
while True:
	keyword = ws.cell(row=(2 + k), column=1).value
	if keyword is None:
		break;
	keyword = str(keyword).strip()

	print("#%d: searching '%s' ..." % (k, keyword))
	items = [list([f[0] for f in fields])]
	j = 0
	while True:
		if anon:
			res = fetch_anon(keyword, j + 1)
		else:
			res = fetch_logged(keyword, j + 1)

		if len(res) < 1:
			break
		items.extend(res)
		j += 1

	k += 1

	filename = outdir + "\\" + keyword + ".xlsx"
	print("writing to file '%s' ..." % filename)
	workbook = xlsxwriter.Workbook(filename)
	worksheet = workbook.add_worksheet()
	for i, item in enumerate(items):
		for j, v in enumerate(item):
			worksheet.write(i, j, v)
	workbook.close()

print("done!")